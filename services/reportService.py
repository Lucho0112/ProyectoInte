import pandas as pd
from datetime import datetime, date, timezone, timedelta
from typing import Dict, List, Callable, Optional
from config.firebaseConfig import firebase_config
from config.settings import Settings
from utils.logger import app_logger, log_audit
from services.firebaseWrapper import requires_connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportService:
    """Servicio para generar y exportar reportes tributarios"""
    
    def __init__(self):
        self.db = firebase_config.get_firestore_client()
        self.datos_ref = self.db.collection(Settings.COLLECTION_DATOS_TRIBUTARIOS)
        self.reportes_ref = self.db.collection(Settings.COLLECTION_REPORTES)
        self.usuarios_ref = self.db.collection(Settings.COLLECTION_USUARIOS)
        self.chile_tz = timezone(timedelta(hours=-3))
        
        # Caché de RUTs para optimizar consultas
        self.rut_cache: Dict[str, str] = {}
        
        # Límites de seguridad
        self.MAX_RECORDS = 10000
        self.MAX_PREVIEW = 50
    
    def get_chile_time(self):
        """Retorna la fecha/hora actual en zona horaria de Chile"""
        return datetime.now(self.chile_tz)
    
    def limpiar_cache(self):
        """Limpia el caché de RUTs"""
        self.rut_cache.clear()
    
    @requires_connection
    def obtener_datos_filtrados(self, filtros: Dict, usuario_id: str, rol: str) -> List[Dict]:
        """
        Obtiene calificaciones tributarias según filtros
        
        Args:
            filtros (Dict): Filtros a aplicar
            usuario_id (str): ID del usuario que genera el reporte
            rol (str): Rol del usuario
            
        Returns:
            List[Dict]: Lista de calificaciones (máximo MAX_RECORDS)
        """
        try:
            # ✅ CAMBIO: Query simple con UN SOLO where() para evitar necesitar índices compuestos
            # Filtramos solo por "activo" en Firestore y el resto en memoria
            query = self.datos_ref.where("activo", "==", True).limit(self.MAX_RECORDS)
            
            # Ejecutar query
            docs = query.stream()
            
            calificaciones = []
            
            # Obtener filtros
            fecha_desde = filtros.get("fecha_desde")
            fecha_hasta = filtros.get("fecha_hasta")
            tipo_impuesto = filtros.get("tipo_impuesto")
            pais = filtros.get("pais")
            rut_cliente_filtro = filtros.get("rut_cliente", "")
            estado_filtro = filtros.get("estado", "ambos")
            
            # ✅ CAMBIO: Filtrar TODO en memoria (Python) en vez de en Firestore
            for doc in docs:
                data = doc.to_dict()
                data["_id"] = doc.id
                
                # Filtro de fechas
                if fecha_desde:
                    fecha_declaracion = data.get("fechaDeclaracion", "")
                    if fecha_declaracion:
                        fecha_str = fecha_desde.strftime("%Y-%m-%d")
                        if fecha_declaracion < fecha_str:
                            continue
                
                if fecha_hasta:
                    fecha_declaracion = data.get("fechaDeclaracion", "")
                    if fecha_declaracion:
                        fecha_str = fecha_hasta.strftime("%Y-%m-%d")
                        if fecha_declaracion > fecha_str:
                            continue
                
                # Filtro de tipo impuesto
                if tipo_impuesto:
                    if data.get("tipoImpuesto", "") != tipo_impuesto:
                        continue
                
                # Filtro de país
                if pais:
                    if data.get("pais", "") != pais:
                        continue
                
                # Filtro de RUT cliente
                if rut_cliente_filtro:
                    cliente_id = data.get("clienteId", "")
                    rut_cliente = self.obtener_rut_cliente(cliente_id)
                    rut_filtro = rut_cliente_filtro.strip().upper()
                    
                    if rut_filtro not in rut_cliente:
                        continue
                
                # Filtrar por estado (local/bolsa)
                es_local = data.get("esLocal", False)
                
                if estado_filtro == "local" and not es_local:
                    continue
                if estado_filtro == "bolsa" and es_local:
                    continue
                
                # Permisos según rol
                es_propietario = data.get("propietarioRegistroId") == usuario_id
                
                if rol == "administrador":
                    calificaciones.append(data)
                elif not es_local:
                    # Datos de bolsa: todos pueden verlos
                    calificaciones.append(data)
                elif es_propietario:
                    # Datos locales: solo el propietario
                    calificaciones.append(data)
            
            app_logger.info(
                f"Usuario {usuario_id[:8]}... obtuvo {len(calificaciones)} registros filtrados"
            )
            
            return calificaciones
        
        except Exception as e:
            app_logger.error(f"Error al obtener datos filtrados: {str(e)}")
            return []
    
    def obtener_rut_cliente(self, cliente_id: str) -> str:
        """
        Obtiene el RUT de un cliente por su ID con caché
        
        Args:
            cliente_id (str): ID del cliente
            
        Returns:
            str: RUT del cliente o "N/A"
        """
        if not cliente_id:
            return "N/A"
        
        # Verificar caché primero
        if cliente_id in self.rut_cache:
            return self.rut_cache[cliente_id]
        
        try:
            doc = self.usuarios_ref.document(cliente_id).get()
            if doc.exists:
                rut = doc.to_dict().get("rut", "N/A")
                # Guardar en caché
                self.rut_cache[cliente_id] = rut
                app_logger.debug(f"RUT cacheado: {cliente_id[:8]}... -> {rut}")
                return rut
            
            # Cliente no existe
            self.rut_cache[cliente_id] = "N/A"
            return "N/A"
        
        except Exception as e:
            app_logger.error(f"Error al obtener RUT de cliente {cliente_id[:8]}...: {e}")
            self.rut_cache[cliente_id] = "N/A"
            return "N/A"
    
    def validar_permisos_exportacion(self, calificaciones: List[Dict], 
                                     usuario_id: str, rol: str) -> List[Dict]:
        """
        Valida que el usuario tenga permisos para exportar los datos
        
        Args:
            calificaciones (List[Dict]): Datos a exportar
            usuario_id (str): ID del usuario
            rol (str): Rol del usuario
            
        Returns:
            List[Dict]: Datos validados
        """
        if rol == "administrador":
            return calificaciones
        
        datos_validos = []
        for dato in calificaciones:
            # Datos de bolsa: todos pueden exportar
            if not dato.get("esLocal", False):
                datos_validos.append(dato)
                continue
            
            # Datos locales: solo el propietario
            if dato.get("propietarioRegistroId") == usuario_id:
                datos_validos.append(dato)
        
        return datos_validos
    
    def preparar_dataframe(self, calificaciones: List[Dict]) -> pd.DataFrame:
        """
        Convierte las calificaciones a DataFrame para exportación
        
        Args:
            calificaciones (List[Dict]): Lista de calificaciones
            
        Returns:
            pd.DataFrame: DataFrame con los datos formateados
        """
        if not calificaciones:
            return pd.DataFrame()
        
        data = []
        
        for idx, cal in enumerate(calificaciones):
            try:
                # Obtener RUT del cliente (usa caché)
                rut_cliente = self.obtener_rut_cliente(cal.get("clienteId", ""))
                
                # ✅ CORRECCIÓN: Manejo robusto de factores
                factores_raw = cal.get("factores", {})
                factores_dict = {}
                
                # Convertir factores a diccionario normalizado
                if isinstance(factores_raw, dict):
                    factores_dict = factores_raw
                elif isinstance(factores_raw, (list, tuple)):
                    # Convertir lista a diccionario
                    for i, valor in enumerate(factores_raw, 1):
                        if i <= 19:
                            factores_dict[f"factor_{i}"] = valor
                else:
                    app_logger.warning(f"Registro {idx}: factores tiene tipo no soportado: {type(factores_raw)}")
                    factores_dict = {}
                
                # Calcular suma factores 8-19
                suma_8_19 = 0.0
                try:
                    suma_8_19 = sum(
                        float(factores_dict.get(f"factor_{i}", 0)) 
                        for i in range(8, 20)
                    )
                except (ValueError, TypeError) as e:
                    app_logger.warning(f"Registro {idx}: Error al sumar factores 8-19: {e}")
                    suma_8_19 = 0.0
                
                # Crear fila con valores SIMPLES (no listas, no objetos complejos)
                fila = {
                    "RUT Cliente": str(rut_cliente),
                    "Fecha Declaración": str(cal.get("fechaDeclaracion", "")),
                    "Tipo Impuesto": str(cal.get("tipoImpuesto", "")),
                    "País": str(cal.get("pais", "")),
                    "Monto Declarado": float(cal.get("montoDeclarado", 0)),
                }
                
                # Agregar factores 1-19 individualmente
                for i in range(1, 20):
                    try:
                        valor = factores_dict.get(f"factor_{i}", 0)
                        # Asegurarse de que sea un número, no una lista
                        if isinstance(valor, (list, tuple)):
                            app_logger.warning(f"Registro {idx}: factor_{i} es una lista, usando primer valor")
                            valor = valor[0] if len(valor) > 0 else 0
                        fila[f"Factor {i}"] = float(valor)
                    except (ValueError, TypeError, IndexError) as e:
                        app_logger.warning(f"Registro {idx}: Error en factor_{i}: {e}")
                        fila[f"Factor {i}"] = 0.0
                
                # Agregar suma y validación
                fila["Suma Factores 8-19"] = float(suma_8_19)
                fila["Estado"] = "Local" if cal.get("esLocal", False) else "Bolsa"
                fila["Válido"] = "Sí" if suma_8_19 <= 1.0 else "No (>1.0)"
                
                data.append(fila)
                
            except Exception as e:
                app_logger.error(f"Error al preparar registro {idx}: {str(e)}")
                import traceback
                app_logger.error(traceback.format_exc())
                continue
        
        # ✅ VALIDACIÓN: Verificar que data no esté vacío
        if not data:
            app_logger.warning("No se pudo preparar ningún registro para el DataFrame")
            return pd.DataFrame()
        
        # Crear DataFrame
        try:
            df = pd.DataFrame(data)
            app_logger.info(f"DataFrame creado exitosamente con {len(df)} filas y {len(df.columns)} columnas")
            return df
        except Exception as e:
            app_logger.error(f"Error al crear DataFrame: {str(e)}")
            import traceback
            app_logger.error(traceback.format_exc())
            return pd.DataFrame()
    
    def exportar_csv(self, file_path: str, calificaciones: List[Dict], 
                    filtros: Dict, usuario_id: str, rol: str,
                    progress_callback: Optional[Callable[[int], None]] = None) -> Dict:
        """
        Exporta calificaciones a CSV
        
        Args:
            file_path (str): Ruta donde guardar el archivo
            calificaciones (List[Dict]): Datos a exportar
            filtros (Dict): Filtros aplicados
            usuario_id (str): ID del usuario
            rol (str): Rol del usuario
            progress_callback: Función para actualizar progreso
            
        Returns:
            Dict: Resultado de la exportación
        """
        try:
            if not calificaciones:
                return {
                    "success": False,
                    "message": "No hay datos para exportar"
                }
            
            # Validar permisos
            calificaciones = self.validar_permisos_exportacion(
                calificaciones, usuario_id, rol
            )
            
            if not calificaciones:
                return {
                    "success": False,
                    "message": "No tiene permisos para exportar estos datos"
                }
            
            if progress_callback:
                progress_callback(20)
            
            # Crear DataFrame
            df = self.preparar_dataframe(calificaciones)
            
            if progress_callback:
                progress_callback(50)
            
            # Exportar a CSV con encoding UTF-8 con BOM para Excel
            df.to_csv(file_path, index=False, encoding='utf-8-sig')
            
            if progress_callback:
                progress_callback(80)
            
            # Registrar en Firebase
            self.registrar_reporte(
                usuario_id=usuario_id,
                tipo_reporte="exportacion_calificaciones",
                filtros=filtros,
                total_registros=len(calificaciones),
                formato="CSV",
                nombre_archivo=file_path.split("/")[-1].split("\\")[-1]
            )
            
            if progress_callback:
                progress_callback(100)
            
            app_logger.info(f"CSV exportado: {file_path} ({len(calificaciones)} registros)")
            
            return {
                "success": True,
                "message": f"CSV generado exitosamente con {len(calificaciones)} registros",
                "file_path": file_path,
                "total_registros": len(calificaciones)
            }
        
        except Exception as e:
            app_logger.error(f"Error al exportar CSV: {str(e)}")
            return {
                "success": False,
                "message": f"Error al exportar: {str(e)}"
            }
    
    def exportar_excel(self, file_path: str, calificaciones: List[Dict], 
                      filtros: Dict, usuario_id: str, rol: str,
                      progress_callback: Optional[Callable[[int], None]] = None) -> Dict:
        """
        Exporta calificaciones a Excel con formato profesional
        
        Args:
            file_path (str): Ruta donde guardar el archivo
            calificaciones (List[Dict]): Datos a exportar
            filtros (Dict): Filtros aplicados
            usuario_id (str): ID del usuario
            rol (str): Rol del usuario
            progress_callback: Función para actualizar progreso
            
        Returns:
            Dict: Resultado de la exportación
        """
        try:
            if not calificaciones:
                return {
                    "success": False,
                    "message": "No hay datos para exportar"
                }
            
            # Validar permisos
            calificaciones = self.validar_permisos_exportacion(
                calificaciones, usuario_id, rol
            )
            
            if not calificaciones:
                return {
                    "success": False,
                    "message": "No tiene permisos para exportar estos datos"
                }
            
            if progress_callback:
                progress_callback(10)
            
            # Crear DataFrame
            df = self.preparar_dataframe(calificaciones)
            
            if progress_callback:
                progress_callback(30)
            
            # Crear workbook
            wb = Workbook()
            
            # === HOJA 1: Datos ===
            ws_datos = wb.active
            ws_datos.title = "Calificaciones Tributarias"
            
            # Colores corporativos
            header_fill = PatternFill(start_color="E94E1B", end_color="E94E1B", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            thin_border = Border(
                left=Side(style='thin', color='E6E9EE'),
                right=Side(style='thin', color='E6E9EE'),
                top=Side(style='thin', color='E6E9EE'),
                bottom=Side(style='thin', color='E6E9EE')
            )
            
            # Escribir datos
            row_idx = 1
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_datos.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border
                    
                    # Formato de encabezado
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        # Formato de datos
                        if c_idx <= 4:  # Texto
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:  # Números
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Resaltar valores inválidos
                        if "Factor" in str(ws_datos.cell(row=1, column=c_idx).value):
                            if isinstance(value, (int, float)) and value > 1.0:
                                cell.fill = PatternFill(
                                    start_color="FFCCCC", 
                                    end_color="FFCCCC", 
                                    fill_type="solid"
                                )
                
                row_idx = r_idx
                
                # Actualizar progreso
                if progress_callback and r_idx % 100 == 0:
                    progress = 30 + int((r_idx / len(df)) * 40)
                    progress_callback(progress)
            
            # Ajustar anchos de columna
            for column in ws_datos.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_datos.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar primera fila
            ws_datos.freeze_panes = "A2"
            
            if progress_callback:
                progress_callback(75)
            
            # === HOJA 2: Resumen ===
            ws_resumen = wb.create_sheet("Resumen")
            
            # Título
            ws_resumen['A1'] = "RESUMEN DE EXPORTACIÓN"
            ws_resumen['A1'].font = Font(size=16, bold=True, color="E94E1B")
            ws_resumen['A1'].alignment = Alignment(horizontal="left")
            
            # Metadatos
            row = 3
            metadata = [
                ("Fecha de generación:", self.get_chile_time().strftime("%Y-%m-%d %H:%M:%S")),
                ("Usuario:", usuario_id[:12] + "..."),
                ("Total de registros:", len(calificaciones)),
                ("", ""),
                ("FILTROS APLICADOS:", ""),
                ("Fecha desde:", filtros.get("fecha_desde", "Sin filtro")),
                ("Fecha hasta:", filtros.get("fecha_hasta", "Sin filtro")),
                ("Tipo de impuesto:", filtros.get("tipo_impuesto", "Todos")),
                ("País:", filtros.get("pais", "Todos")),
                ("Estado:", filtros.get("estado", "Ambos")),
            ]
            
            for label, value in metadata:
                ws_resumen[f'A{row}'] = label
                ws_resumen[f'A{row}'].font = Font(bold=True)
                ws_resumen[f'B{row}'] = str(value)
                row += 1
            
            # Estadísticas
            row += 1
            ws_resumen[f'A{row}'] = "ESTADÍSTICAS"
            ws_resumen[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            # Calcular estadísticas
            locales = sum(1 for c in calificaciones if c.get("esLocal", False))
            bolsa = len(calificaciones) - locales
            total_monto = sum(c.get("montoDeclarado", 0) for c in calificaciones)
            
            # Contar registros válidos/inválidos
            validos = 0
            invalidos = 0
            for c in calificaciones:
                factores = c.get("factores", {})
                suma = sum(factores.get(f"factor_{i}", 0) for i in range(8, 20))
                if suma <= 1.0:
                    validos += 1
                else:
                    invalidos += 1
            
            stats = [
                ("Registros Locales:", locales),
                ("Registros Bolsa:", bolsa),
                ("Registros Válidos:", validos),
                ("Registros Inválidos:", invalidos),
                ("Monto Total Declarado:", f"${total_monto:,.2f}"),
                ("Monto Promedio:", f"${total_monto/len(calificaciones):,.2f}" if calificaciones else "$0.00"),
            ]
            
            for label, value in stats:
                ws_resumen[f'A{row}'] = label
                ws_resumen[f'A{row}'].font = Font(bold=True)
                ws_resumen[f'B{row}'] = value
                row += 1
            
            # Ajustar anchos
            ws_resumen.column_dimensions['A'].width = 30
            ws_resumen.column_dimensions['B'].width = 40
            
            if progress_callback:
                progress_callback(90)
            
            # Guardar
            wb.save(file_path)
            
            if progress_callback:
                progress_callback(95)
            
            # Registrar en Firebase
            self.registrar_reporte(
                usuario_id=usuario_id,
                tipo_reporte="exportacion_calificaciones",
                filtros=filtros,
                total_registros=len(calificaciones),
                formato="Excel",
                nombre_archivo=file_path.split("/")[-1].split("\\")[-1]
            )
            
            if progress_callback:
                progress_callback(100)
            
            app_logger.info(f"Excel exportado: {file_path} ({len(calificaciones)} registros)")
            
            return {
                "success": True,
                "message": f"Excel generado exitosamente con {len(calificaciones)} registros",
                "file_path": file_path,
                "total_registros": len(calificaciones)
            }
        
        except Exception as e:
            app_logger.error(f"Error al exportar Excel: {str(e)}")
            return {
                "success": False,
                "message": f"Error al exportar: {str(e)}"
            }
    
    @requires_connection
    def registrar_reporte(self, usuario_id: str, tipo_reporte: str, 
                         filtros: Dict, total_registros: int, 
                         formato: str, nombre_archivo: str) -> bool:
        """
        Registra un reporte generado en Firebase
        
        Args:
            usuario_id (str): ID del usuario que generó el reporte
            tipo_reporte (str): Tipo de reporte
            filtros (Dict): Filtros aplicados
            total_registros (int): Total de registros exportados
            formato (str): Formato del reporte (CSV/Excel)
            nombre_archivo (str): Nombre del archivo generado
            
        Returns:
            bool: True si se registró correctamente
        """
        try:
            # ✅ CORRECCIÓN: Convertir tipos de datos para Firestore
            filtros_firestore = {}
            for key, value in filtros.items():
                if isinstance(value, date) and not isinstance(value, datetime):
                    # Convertir date a datetime
                    filtros_firestore[key] = datetime.combine(value, datetime.min.time())
                elif isinstance(value, (list, dict, set, tuple)):
                    # Convertir tipos complejos a string para evitar error "unhashable type"
                    filtros_firestore[key] = str(value)
                elif value is None:
                    # Firestore acepta None
                    filtros_firestore[key] = value
                elif isinstance(value, (str, int, float, bool)):
                    # Tipos básicos soportados por Firestore
                    filtros_firestore[key] = value
                else:
                    # Cualquier otro tipo, convertir a string por seguridad
                    filtros_firestore[key] = str(value)
            
            # Crear diccionario del reporte
            reporte_data = {
                "usuarioGeneradorId": usuario_id,
                "tipoReporte": tipo_reporte,
                "filtrosAplicados": filtros_firestore,
                "totalRegistros": total_registros,
                "formato": formato,
                "nombreArchivo": nombre_archivo,
                "fechaGeneracion": self.get_chile_time()
            }
            
            # Guardar en Firestore
            self.reportes_ref.add(reporte_data)
            
            # Registrar auditoría
            log_audit(
                action="REPORTE_GENERADO",
                user_id=usuario_id,
                details={
                    "tipo": tipo_reporte,
                    "formato": formato,
                    "registros": total_registros,
                    "archivo": nombre_archivo
                }
            )
            
            app_logger.info(f"Reporte registrado: {nombre_archivo} por usuario {usuario_id[:8]}...")
            return True
        
        except Exception as e:
            app_logger.error(f"Error al registrar reporte: {str(e)}")
            return False
    
    @requires_connection
    def obtener_historial_reportes(self, usuario_id: str, rol: str) -> List[Dict]:
        """
        Obtiene el historial de reportes generados
        
        Args:
            usuario_id (str): ID del usuario
            rol (str): Rol del usuario
            
        Returns:
            List[Dict]: Lista de reportes (últimos 50)
        """
        try:
            # Admin/Auditor ven todos, otros usuarios ven solo los suyos
            if rol in ["administrador", "auditor_tributario"]:
                query = self.reportes_ref\
                    .order_by("fechaGeneracion", direction="DESCENDING")\
                    .limit(50)
            else:
                query = self.reportes_ref\
                    .where("usuarioGeneradorId", "==", usuario_id)\
                    .order_by("fechaGeneracion", direction="DESCENDING")\
                    .limit(50)
            
            docs = query.stream()
            
            reportes = []
            for doc in docs:
                data = doc.to_dict()
                data["_id"] = doc.id
                reportes.append(data)
            
            app_logger.info(f"Historial cargado: {len(reportes)} reportes para usuario {usuario_id[:8]}...")
            return reportes
        
        except Exception as e:
            app_logger.error(f"Error al obtener historial: {str(e)}")
            return []